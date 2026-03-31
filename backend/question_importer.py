#!/usr/bin/env python3
"""
题库导入脚本 - 将Excel题库文件导入到密评备考系统
"""

import openpyxl
import json
import psycopg2
import sys
import os

class QuestionImporter:
    def __init__(self, db_config):
        self.db_config = db_config
        self.conn = None
        
    def connect_db(self):
        """连接到数据库"""
        try:
            self.conn = psycopg2.connect(**self.db_config)
            return True
        except Exception as e:
            print(f"数据库连接失败: {e}")
            return False
            
    def create_tables(self):
        """创建数据库表"""
        try:
            cursor = self.conn.cursor()
            
            # 创建题目表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS questions (
                    id SERIAL PRIMARY KEY,
                    question_no INTEGER NOT NULL,
                    question_type VARCHAR(50) NOT NULL,
                    question_text TEXT NOT NULL,
                    option_a TEXT,
                    option_b TEXT,
                    option_c TEXT,
                    option_d TEXT,
                    correct_answer VARCHAR(10) NOT NULL,
                    original_no VARCHAR(50),
                    difficulty_level INTEGER DEFAULT 1,
                    knowledge_point VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 创建知识点表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS knowledge_points (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(100) NOT NULL UNIQUE,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            self.conn.commit()
            print("✅ 数据库表创建成功")
            return True
            
        except Exception as e:
            print(f"创建表失败: {e}")
            return False
            
    def import_excel(self, file_path):
        """导入Excel题库文件"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 统计信息
            total_rows = ws.max_row - 2  # 减去标题行
            imported_count = 0
            
            cursor = self.conn.cursor()
            
            for row in range(3, ws.max_row + 1):  # 从第3行开始（跳过标题）
                try:
                    # 读取数据
                    question_no = ws.cell(row=row, column=1).value or 0
                    question_type = ws.cell(row=row, column=2).value or ""
                    question_text = ws.cell(row=row, column=3).value or ""
                    option_a = ws.cell(row=row, column=4).value or ""
                    option_b = ws.cell(row=row, column=5).value or ""
                    option_c = ws.cell(row=row, column=6).value or ""
                    option_d = ws.cell(row=row, column=7).value or ""
                    correct_answer = ws.cell(row=row, column=8).value or ""
                    original_no = ws.cell(row=row, column=9).value or ""
                    
                    # 插入数据库
                    cursor.execute("""
                        INSERT INTO questions 
                        (question_no, question_type, question_text, option_a, option_b, option_c, option_d, correct_answer, original_no)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (question_no, question_type, question_text, option_a, option_b, option_c, option_d, correct_answer, original_no))
                    
                    imported_count += 1
                    
                    if imported_count % 100 == 0:
                        print(f"已导入 {imported_count}/{total_rows} 条题目")
                        
                except Exception as e:
                    print(f"导入第{row}行失败: {e}")
                    continue
            
            self.conn.commit()
            print(f"✅ 题库导入完成！共导入 {imported_count} 条题目")
            return imported_count
            
        except Exception as e:
            print(f"导入Excel文件失败: {e}")
            return 0
            
    def get_statistics(self):
        """获取题库统计信息"""
        try:
            cursor = self.conn.cursor()
            
            # 总题目数
            cursor.execute("SELECT COUNT(*) FROM questions")
            total_questions = cursor.fetchone()[0]
            
            # 题型分布
            cursor.execute("SELECT question_type, COUNT(*) FROM questions GROUP BY question_type")
            type_distribution = cursor.fetchall()
            
            print(f"\n📊 题库统计信息:")
            print(f"总题目数: {total_questions}")
            print(f"题型分布:")
            for q_type, count in type_distribution:
                print(f"  {q_type}: {count}题")
                
            return total_questions, type_distribution
            
        except Exception as e:
            print(f"获取统计信息失败: {e}")
            return 0, []

def main():
    """主函数"""
    # 数据库配置
    db_config = {
        'host': 'localhost',
        'port': 15432,
        'database': 'exam_db',
        'user': 'exam_user',
        'password': 'exam_pass123'
    }
    
    # Excel文件路径
    excel_file = '/home/pi/.openclaw/media/inbound/å_ç_å_ç_åº_ç_å_å_æ_è_ä¼_ä_ä_äººå_è_æ_é_åº_-å_è_ç_æ---94da4053-efc3-42e2-917c-699acbea321a.xlsx'
    
    # 创建导入器
    importer = QuestionImporter(db_config)
    
    # 连接数据库
    if not importer.connect_db():
        print("❌ 数据库连接失败，请检查数据库服务是否运行")
        return
        
    # 创建表
    if not importer.create_tables():
        print("❌ 创建数据库表失败")
        return
        
    # 导入Excel数据
    print("🚀 开始导入题库数据...")
    imported_count = importer.import_excel(excel_file)
    
    if imported_count > 0:
        # 获取统计信息
        importer.get_statistics()
        print("\n🎉 题库导入完成！系统已准备好使用。")
    else:
        print("❌ 题库导入失败")
        
    # 关闭连接
    if importer.conn:
        importer.conn.close()

if __name__ == "__main__":
    main()