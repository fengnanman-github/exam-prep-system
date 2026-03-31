#!/usr/bin/env python3
"""
简化的题库导入脚本
"""

import openpyxl
import psycopg2
import sys

def import_questions():
    # 数据库配置
    db_config = {
        'host': 'localhost',
        'port': 15432,
        'database': 'exam_db',
        'user': 'exam_user',
        'password': 'exam_pass123'
    }
    
    # Excel文件路径
    excel_file = '/home/pi/.openclaw/media/inbound/å_ç_å_ç_åº_ç_å_å_æ_è_ä¼_ä_ä_äººå_è_æ_é_åº_-å_è_ç_æ---94da4053-efc3-42e2-917c-699acbea321a.xlsx'
    
    try:
        # 连接数据库
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"工作表: {ws.title}")
        print(f"数据范围: A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
        print(f"总行数: {ws.max_row}")
        
        # 清空现有数据
        cursor.execute("DELETE FROM questions")
        
        imported_count = 0
        error_count = 0
        
        # 从第3行开始导入（跳过标题）
        for row in range(3, ws.max_row + 1):
            try:
                # 读取数据
                question_no = ws.cell(row=row, column=1).value or 0
                question_type = ws.cell(row=row, column=2).value or ""
                question_text = ws.cell(row=row, column=3).value or ""
                option_a = ws.cell(row=row, column=4).value or ""
                option_b = ws.cell(row=row, column=5).value or ""
                option_c = ws.cell(row=row, column=6).value or ""
                option_d = ws.cell(row=row, column=7).value or ""
                correct_answer = ws.cell(row=row, column=8).value or ""
                original_no = ws.cell(row=row, column=9).value or ""
                
                # 跳过空行
                if not question_text:
                    continue
                
                # 插入数据库
                cursor.execute("""
                    INSERT INTO questions 
                    (question_no, question_type, question_text, option_a, option_b, option_c, option_d, correct_answer, original_no)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (question_no, question_type, question_text, option_a, option_b, option_c, option_d, correct_answer, original_no))
                
                imported_count += 1
                
                if imported_count % 100 == 0:
                    print(f"已导入 {imported_count} 条题目")
                    
            except Exception as e:
                error_count += 1
                if error_count <= 5:  # 只显示前5个错误
                    print(f"第{row}行导入失败: {e}")
                continue
        
        conn.commit()
        
        print(f"\n✅ 导入完成！")
        print(f"成功导入: {imported_count} 条题目")
        print(f"导入失败: {error_count} 条题目")
        
        # 统计信息
        cursor.execute("SELECT COUNT(*) FROM questions")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT question_type, COUNT(*) FROM questions GROUP BY question_type")
        types = cursor.fetchall()
        
        print(f"\n📊 题库统计:")
        print(f"总题目数: {total}")
        print("题型分布:")
        for t, c in types:
            print(f"  {t}: {c}题")
        
        conn.close()
        
    except Exception as e:
        print(f"❌ 导入失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import_questions()