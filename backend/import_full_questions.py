#!/usr/bin/env python3
"""
完整题库导入脚本
在Docker容器中运行
"""

# 首先安装必要的库
import subprocess
import sys

print("安装Python库...")
subprocess.run(["apk", "add", "--no-cache", "py3-pip"], check=False)
subprocess.run(["pip3", "install", "openpyxl", "pandas"], check=False)

import psycopg2
from openpyxl import load_workbook
import re

# 数据库配置
db_config = {
    'host': 'db',  # Docker内部网络
    'port': 5432,
    'database': 'exam_db',
    'user': 'exam_user',
    'password': 'exam_pass123'
}

excel_file = '/app/商用密码应用安全性评估从业人员考核题库-参考答案.xlsx'

def categorize_question(question_text, options_text=""):
    """根据题目内容自动分类"""
    text = (question_text + " " + options_text).lower()

    # 密码算法相关
    algorithm_keywords = ['算法', 'sm2', 'sm3', 'sm4', 'sm9', 'aes', 'des', 'rsa', 'ecc', '哈希', '摘要', '加密', '解密', '对称', '非对称', '密钥']
    # 密码协议相关
    protocol_keywords = ['协议', 'ssl', 'tls', 'ssh', 'ipsec', 'https', 'vpn', '握手', '通信', '传输']
    # PKI体系相关
    pki_keywords = ['pki', '证书', 'ca', '认证', '公钥', '私钥', '数字证书', '证书授权', '签发', '吊销', 'crl']
    # 密码产品相关
    product_keywords = ['产品', '密码机', '网关', '服务器', '模块', '芯片', '密码设备', '审批', '销售', '进口']
    # 密码管理相关
    management_keywords = ['管理', '策略', '制度', '人员', '培训', '监督', '检查', '评估', '审计', '责任', '密码管理']

    scores = {
        '密码算法': sum(1 for kw in algorithm_keywords if kw in text),
        '密码协议': sum(1 for kw in protocol_keywords if kw in text),
        'PKI体系': sum(1 for kw in pki_keywords if kw in text),
        '密码产品': sum(1 for kw in product_keywords if kw in text),
        '密码管理': sum(1 for kw in management_keywords if kw in text)
    }

    # 返回得分最高的分类
    max_category = max(scores, key=scores.get)
    if scores[max_category] == 0:
        return '密码管理'  # 默认分类
    return max_category

def determine_difficulty(category):
    """根据分类确定难度（简化版）"""
    # 这里可以根据更复杂的逻辑来确定难度
    import random
    return random.choice(['easy', 'medium', 'medium', 'hard'])

def import_questions():
    try:
        print(f"正在读取Excel文件: {excel_file}")
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        print(f"工作表: {ws.title}")
        print(f"总行数: {ws.max_row}")

        # 分析表头
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)

        print(f"\n列名: {headers}")

        # 找到关键列的索引
        col_mapping = {}
        for i, h in enumerate(headers):
            if h:
                h_lower = str(h).lower()
                if '题号' in h_lower or '编号' in h_lower:
                    col_mapping['question_no'] = i
                elif '题型' in h_lower or '类型' in h_lower:
                    col_mapping['question_type'] = i
                elif '题目' in h_lower or '问题' in h_lower or '内容' in h_lower:
                    col_mapping['question_text'] = i
                elif 'a选项' in h_lower or h_lower == 'a':
                    col_mapping['option_a'] = i
                elif 'b选项' in h_lower or h_lower == 'b':
                    col_mapping['option_b'] = i
                elif 'c选项' in h_lower or h_lower == 'c':
                    col_mapping['option_c'] = i
                elif 'd选项' in h_lower or h_lower == 'd':
                    col_mapping['option_d'] = i
                elif '答案' in h_lower or '参考答案' in h_lower:
                    col_mapping['correct_answer'] = i
                elif '解析' in h_lower or '说明' in h_lower:
                    col_mapping['explanation'] = i
                elif '知识点' in h_lower:
                    col_mapping['knowledge_point'] = i

        print(f"\n列映射: {col_mapping}")

        if 'question_text' not in col_mapping:
            print("错误: 找不到题目列！")
            return

        # 连接数据库
        print("\n连接数据库...")
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()

        # 清空现有数据
        print("清空现有题库...")
        cursor.execute("DELETE FROM questions")
        conn.commit()

        # 导入数据
        imported = 0
        errors = 0
        category_stats = {}

        print(f"\n开始导入（从第2行到第{ws.max_row}行）...")

        for row in range(2, ws.max_row + 1):
            try:
                # 读取题目内容
                q_text = ws.cell(row=row, column=col_mapping['question_text'] + 1).value

                if not q_text or str(q_text).strip() == '':
                    continue

                # 读取其他字段
                q_type = ws.cell(row=row, column=col_mapping.get('question_type', 1) + 1).value if 'question_type' in col_mapping else '判断题'
                opt_a = ws.cell(row=row, column=col_mapping.get('option_a', 3) + 1).value if 'option_a' in col_mapping else None
                opt_b = ws.cell(row=row, column=col_mapping.get('option_b', 4) + 1).value if 'option_b' in col_mapping else None
                opt_c = ws.cell(row=row, column=col_mapping.get('option_c', 5) + 1).value if 'option_c' in col_mapping else None
                opt_d = ws.cell(row=row, column=col_mapping.get('option_d', 6) + 1).value if 'option_d' in col_mapping else None
                answer = ws.cell(row=row, column=col_mapping.get('correct_answer', 7) + 1).value if 'correct_answer' in col_mapping else 'A'
                knowledge = ws.cell(row=row, column=col_mapping.get('knowledge_point', 8) + 1).value if 'knowledge_point' in col_mapping else None
                explanation = ws.cell(row=row, column=col_mapping.get('explanation', 9) + 1).value if 'explanation' in col_mapping else None

                # 清理数据
                q_text = str(q_text).strip()
                q_type = str(q_type).strip() if q_type else '判断题'
                answer = str(answer).strip().upper() if answer else 'A'

                # 判断题型
                if '判断' in q_type or q_type in ['对', '错', '正确', '错误']:
                    q_type = '判断题'
                    if not opt_a:
                        opt_a = '正确'
                    if not opt_b:
                        opt_b = '错误'
                    # 标准化答案
                    if answer in ['对', '√', '✓', '正确', 'T', 'TRUE']:
                        answer = 'A'
                    elif answer in ['错', '×', '✗', '错误', 'F', 'FALSE']:
                        answer = 'B'
                elif '单选' in q_type or q_type == '单项选择题':
                    q_type = '单项选择题'
                elif '多选' in q_type or q_type == '多项选择题':
                    q_type = '多项选择题'
                else:
                    # 根据选项判断
                    if opt_c and opt_d:
                        q_type = '多项选择题'
                    else:
                        q_type = '单项选择题'

                # 智能分类
                options_text = f"{opt_a or ''} {opt_b or ''} {opt_c or ''} {opt_d or ''}"
                category = categorize_question(q_text, options_text)
                difficulty = determine_difficulty(category)

                # 插入数据库
                cursor.execute("""
                    INSERT INTO questions
                    (question_no, question_type, question_text, option_a, option_b, option_c, option_d,
                     correct_answer, category, difficulty, knowledge_point, explanation)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    imported + 1,  # 临时题号
                    q_type,
                    q_text,
                    str(opt_a).strip() if opt_a else None,
                    str(opt_b).strip() if opt_b else None,
                    str(opt_c).strip() if opt_c else None,
                    str(opt_d).strip() if opt_d else None,
                    answer,
                    category,
                    difficulty,
                    str(knowledge).strip() if knowledge else None,
                    str(explanation).strip() if explanation else None
                ))

                imported += 1
                category_stats[category] = category_stats.get(category, 0) + 1

                # 每100条提交一次
                if imported % 100 == 0:
                    conn.commit()
                    print(f"已导入 {imported} 题...")

            except Exception as e:
                errors += 1
                if errors <= 10:
                    print(f"第{row}行导入失败: {e}")
                continue

        # 最终提交
        conn.commit()

        print(f"\n{'='*60}")
        print(f"导入完成！")
        print(f"{'='*60}")
        print(f"成功导入: {imported} 题")
        print(f"导入失败: {errors} 题")
        print(f"\n分类统计:")
        for cat, count in sorted(category_stats.items()):
            print(f"  {cat}: {count}题")

        # 更新题号
        print(f"\n更新题号...")
        cursor.execute("UPDATE questions SET question_no = id")
        conn.commit()

        # 统计题型
        cursor.execute("SELECT question_type, COUNT(*) FROM questions GROUP BY question_type")
        type_stats = cursor.fetchall()
        print(f"\n题型分布:")
        for qtype, count in type_stats:
            print(f"  {qtype}: {count}题")

        conn.close()
        print(f"\n✅ 导入成功！")

    except Exception as e:
        print(f"❌ 导入失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import_questions()
